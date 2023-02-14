import tkinter
import tkinter.messagebox
import customtkinter
import pyexcel as p
import pandas as pd
import math
from openpyxl import Workbook, load_workbook
from tkinter import filedialog
from tkinter import Toplevel

customtkinter.set_appearance_mode("Dark")
customtkinter.set_default_color_theme("blue")

global count
global sheet_names
sheet_names = ['Thứ 2 - Món 1','Thứ 2 - Món 2','Thứ 3 - Món 1','Thứ 3 - Món 2',
                      'Thứ 4 - Món 1','Thứ 4 - Món 2','Thứ 5 - Món 1','Thứ 5 - Món 2',
                      'Thứ 6-Món 1','Thứ 6-Món 2','Thứ 7-Món 1','Thứ 7- Món 2',
                      'Chủ nhật - Món 1','Chủ nhật -Món 2']
global special_labels
special_labels = [" Muối (Salt)", " Nước tương",
                  " Nước mắm loại I (Fish sauce, liquid, category I)",
                  "Bánh Karo phô mai hoàng kim (26g/cái)",
                  " Nước tương", "Hạt nêm Knorr 100g"]
global salty_labels
salty_labels = [" Nước tương", " Nước mắm loại I (Fish sauce, liquid, category I)"]
count = 3
waste_wb = load_workbook(filename='waste.xlsx')
sh_waste = waste_wb['Sheet1']

class App(customtkinter.CTk):
    
    class popup_done(customtkinter.CTk):
        def __init__(self):
            super().__init__()
            self.title("Promt")
            self.geometry("210x100")
            self.protocol("WM_DELETE_WINDOW", self.on_closing)

            # ============ create_frames ============
            self.grid_columnconfigure(1, weight=1)
            self.grid_rowconfigure(0, weight=1)
            self.frame = customtkinter.CTkFrame(master=self)
            self.frame.grid(row=0, column=1, sticky="nswe", padx=20, pady=20)

            self.label_promt = customtkinter.CTkLabel(master=self.frame,text="DONE")
            self.label_promt.grid(row=0, column=0, columnspan=1, pady=5, padx=5, sticky="")

        def on_closing(self, event=0):
            self.destroy()
            
    class popup_failed(customtkinter.CTk):
        def __init__(self):
            super().__init__()
            self.title("Promt")
            self.geometry("210x100")
            self.protocol("WM_DELETE_WINDOW", self.on_closing)

            # ============ create_frames ============
            self.grid_columnconfigure(1, weight=1)
            self.grid_rowconfigure(0, weight=1)
            self.frame = customtkinter.CTkFrame(master=self)
            self.frame.grid(row=0, column=1, sticky="nswe", padx=20, pady=20)

            self.label_promt = customtkinter.CTkLabel(master=self.frame,text="FAILED")
            self.label_promt.grid(row=0, column=0, columnspan=1, pady=5, padx=5, sticky="")

        def on_closing(self, event=0):
            self.destroy()
    
    WIDTH = 1300
    HEIGHT = 350
    
    def __init__(self):
        
        super().__init__()

        self.title("Eiyokun Helper - CIH")
        self.geometry(f"{App.WIDTH}x{App.HEIGHT}")
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

        # ============ create_frames ============
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.frame = customtkinter.CTkFrame(master=self)
        self.frame.grid(row=0, column=1, sticky="nswe", padx=20, pady=20)

        # ============ frame_main ===============
        self.label_BT01 = customtkinter.CTkLabel(master=self.frame,text="BT01")
        self.label_DD01 = customtkinter.CTkLabel(master=self.frame,text="DD01")
        self.label_TM01 = customtkinter.CTkLabel(master=self.frame,text="TM01")

        self.label_BT01.grid(row=0, column=0, columnspan=1, pady=5, padx=5, sticky="")
        self.label_DD01.grid(row=1, column=0, columnspan=1, pady=5, padx=5, sticky="")
        self.label_TM01.grid(row=2, column=0, columnspan=1, pady=5, padx=5, sticky="")
        
        self.label_BT01_info = customtkinter.CTkLabel(master=self.frame, text="BT01",
                                                      height=15,
                                                      corner_radius=6,
                                                      fg_color=("white", "gray38"),
                                                      justify=tkinter.LEFT)
        self.label_DD01_info = customtkinter.CTkLabel(master=self.frame, text="DD01",
                                                      height=15,
                                                      corner_radius=6,
                                                      fg_color=("white", "gray38"),
                                                      justify=tkinter.LEFT)
        self.label_TM01_info = customtkinter.CTkLabel(master=self.frame, text="TM01",
                                                      height=15,
                                                      corner_radius=6,
                                                      fg_color=("white", "gray38"),
                                                      justify=tkinter.LEFT)
        
        self.label_BT01_info.grid(column=1, row=0, sticky="", padx=10, pady=15)
        self.label_DD01_info.grid(column=1, row=1, sticky="", padx=10, pady=10)
        self.label_TM01_info.grid(column=1, row=2, sticky="", padx=10, pady=10)

        self.button_BT01_path = customtkinter.CTkButton(master=self.frame,
                                                        text="BT01",
                                                        command=self.button_event_BT01_path)
        self.button_DD01_path = customtkinter.CTkButton(master=self.frame,
                                                        text="DD01",
                                                        command=self.button_event_DD01_path)
        self.button_TM01_path = customtkinter.CTkButton(master=self.frame,
                                                        text="TM01",
                                                        command=self.button_event_TM01_path)
        
        self.button_BT01_path.grid(column=3, row=0, padx=10, pady=10)
        self.button_DD01_path.grid(column=3, row=1, padx=10, pady=10)
        self.button_TM01_path.grid(column=3, row=2, padx=10, pady=10)

        self.button_start= customtkinter.CTkButton(master=self.frame,
                                                        text="[1] START BT01 MAIN",
                                                        command=self.button_start_BT01)
        self.button_start.grid(column=0, row=4, padx=10, pady=10)

        self.button_start= customtkinter.CTkButton(master=self.frame,
                                                        text="[2] ADD ALL (DD01,TM01)",
                                                        command=self.button_add_all)
        self.button_start.grid(column=1, row=4, padx=10, pady=10)
        
        self.button_start= customtkinter.CTkButton(master=self.frame,
                                                        text="[2.a] ADD DD01",
                                                        command=self.button_add_DD01)
        self.button_start.grid(column=0, row=5, padx=10, pady=10)
        
        self.button_start= customtkinter.CTkButton(master=self.frame,
                                                        text="[2.b] ADD TM01",
                                                        command=self.button_add_TM01)
        self.button_start.grid(column=1, row=5, padx=10, pady=10)

        # set default values #BT01 DD01 TM01 BT02

        
    def button_event_BT01_path(self):
        global filename_BT01
        filename_BT01 = filedialog.askopenfilename(initialdir="/",title="Chon file BT01 (.xls)", filetypes=(("Microsoft Excel 97-2003 Worksheet (.xls)","*.xls"),("All files","*.*")) )
        self.label_BT01_info = customtkinter.CTkLabel(master=self.frame, text=filename_BT01,
                                                      height=15,
                                                      corner_radius=6,
                                                      fg_color=("white", "gray38"),
                                                      justify=tkinter.LEFT)
        self.label_BT01_info.grid(column=1, row=0, sticky="", padx=10, pady=15)
    
    def button_event_DD01_path(self):
        global filename_DD01
        filename_DD01 = filedialog.askopenfilename(initialdir="/",title="Chon file DD01 (.xls)", filetypes=(("Microsoft Excel 97-2003 Worksheet (.xls)","*.xls"),("All files","*.*")) )
        self.label_DD01_info = customtkinter.CTkLabel(master=self.frame, text=filename_DD01,
                                                      height=15,
                                                      corner_radius=6,
                                                      fg_color=("white", "gray38"),
                                                      justify=tkinter.LEFT)
        self.label_DD01_info.grid(column=1, row=1, sticky="", padx=10, pady=10)
        
    def button_event_TM01_path(self):
        global filename_TM01
        filename_TM01 = filedialog.askopenfilename(initialdir="/",title="Chon file TM01 (.xls)", filetypes=(("Microsoft Excel 97-2003 Worksheet (.xls)","*.xls"),("All files","*.*")) )
        self.label_TM01_info = customtkinter.CTkLabel(master=self.frame, text=filename_TM01,
                                                      height=15,
                                                      corner_radius=6,
                                                      fg_color=("white", "gray38"),
                                                      justify=tkinter.LEFT)
        self.label_TM01_info.grid(column=1, row=2, sticky="", padx=10, pady=10)

    def button_start_BT01(self):
        def roundup(n):
            n = math.trunc(n)
            if (n < 5):
                return n
            elif (n % 10 > 5):
                while n % 10 != 0:
                    n+=1
                return n
            elif (n % 10 < 5):
                while n % 10 != 0:
                    n-=1
                return n
            else:
                return n
        try:
            ###############################
            p.save_book_as(file_name=filename_BT01,
                           dest_file_name='demo.xlsx')
            ###############################
            switch = 0
            workbook = Workbook()
            workbook['Sheet'].title = 'Thứ 2 - Món 1'
            for name in sheet_names:
                workbook.create_sheet(name)

            def write_70717(i):
                sheet.cell(row=i,column=4).value = sh_demo.cell(row=i, column=5).value - 1
                sheet.cell(row=i,column=5).value = 0
                sheet.cell(row=i,column=6).value = sh_demo.cell(row=i, column=5).value - 1
                sheet.cell(row=i,column=7).value = 1
                sheet.cell(row=i,column=8).value = sh_demo.cell(row=i, column=5).value - 1
                sheet.cell(row=i-1,column=4).value = 7
                sheet.cell(row=i-1,column=5).value = 0
                sheet.cell(row=i-1,column=6).value = 7
                sheet.cell(row=i-1,column=7).value = 1
                sheet.cell(row=i-1,column=8).value = 7
                
            def write_it(sheet, sh_demo, row_ct, col_ct):
                switch = 0
                for i in range(1, row_ct+1):
                    for j in range(3,6):          
                        sheet.cell(row=i, column=j-1).value = sh_demo.cell(row=i, column=j).value
                        sheet.cell(row=i, column=1).value = sh_demo.cell(row=i, column=1).value
                        if (switch == 1):
                            sheet.cell(row=i, column=j-1).value = ''
                            sheet.cell(row=i, column=j-2).value = ''
                            switch = 0
                        if (type(sheet.cell(row=i, column=j-1).value) is str):
                            if (sheet.cell(row=i, column=j-1).value[:3] == 'SUM'):
                                switch = 1

                switch = 0
                for i in range(3, row_ct):
                    for j in range(1, sh_waste.max_row):
                        if (sh_demo.cell(row=i, column=4).value == sh_waste.cell(row=j, column=2).value):
                            try:
                                if switch == 1:
                                    switch = 0
                                    print('switch')
                                    pass
                                elif sheet.cell(row=i,column=3).value not in special_labels:
                                    sheet.cell(row=i,column=5).value = sh_waste.cell(row=j, column=4).value
                                    sheet.cell(row=i,column=6).value = roundup(((float(sh_waste.cell(row=j, column=4).value) + 100) * sh_demo.cell(row=i, column=5).value)/100)
                                    sheet.cell(row=i,column=7).value = sh_waste.cell(row=j, column=5).value
                                    sheet.cell(row=i,column=8).value = roundup(float(sh_demo.cell(row=i, column=5).value/float(sh_waste.cell(row=j, column=5).value)))
                                elif (sheet.cell(row=i,column=3).value == " Muối (Salt)"):
                                    sheet.cell(row=i,column=5).value = sh_waste.cell(row=j, column=4).value
                                    sheet.cell(row=i,column=6).value = ((float(sh_waste.cell(row=j, column=4).value) + 100) * sh_demo.cell(row=i, column=5).value)/100
                                    sheet.cell(row=i,column=7).value = sh_waste.cell(row=j, column=5).value
                                    sheet.cell(row=i,column=8).value = float(sh_demo.cell(row=i, column=5).value/float(sh_waste.cell(row=j, column=5).value))
                                    if sheet.cell(row=i-1,column=3).value in salty_labels:
                                        if (sheet.cell(row=i-1,column=4).value == 0) and (sheet.cell(row=i,column=3).value > 1):
                                            write_70717(i)
                                    elif sheet.cell(row=i+1,column=3).value in salty_labels:
                                        if (sheet.cell(row=i+1,column=4).value == 0) and (sheet.cell(row=i,column=3).value > 1):
                                            write_70717(i)
                                            switch = 1
                                else:
                                    sheet.cell(row=i,column=5).value = sh_waste.cell(row=j, column=4).value
                                    sheet.cell(row=i,column=6).value = ((float(sh_waste.cell(row=j, column=4).value) + 100) * sh_demo.cell(row=i, column=5).value)/100
                                    sheet.cell(row=i,column=7).value = sh_waste.cell(row=j, column=5).value
                                    sheet.cell(row=i,column=8).value = float(sh_demo.cell(row=i, column=5).value/float(sh_waste.cell(row=j, column=5).value))
                                break
                            except:
                                pass
                    
                    if (sheet.cell(row=i, column=4).value == "[Weight]"):
                        sheet.cell(row=i, column=4).value = "BT01"
                        sheet.cell(row=i, column=5).value = "[Waste]"
                        sheet.cell(row=i, column=6).value = "BT01"
                        sheet.cell(row=i, column=7).value = "[Tỉ lệ sống/chín]"
                        sheet.cell(row=i, column=8).value = "BT01"

                        sheet.cell(row=i+1, column=5).value = "(%)"
                        sheet.cell(row=i+1, column=6).value = "(g)"
                        sheet.cell(row=i+1, column=7).value = "(%)"
                        sheet.cell(row=i+1, column=8).value = "(g)"

                sheet.cell(row=1, column=4).value = "[Weight]"
                sheet.cell(row=1, column=5).value = "[Waste]"
                sheet.cell(row=1, column=6).value = "[Đi chợ]"
                sheet.cell(row=1, column=7).value = "[Tỉ lệ sống/chín]"
                sheet.cell(row=1, column=8).value = "[Phân chia]"

                  
            ################Thứ 2 - Món 1#########################
            sheet = workbook['Thứ 2 - Món 1']
            wb = load_workbook(filename='demo.xlsx')
            sh_demo = wb['Thứ 2 - Món 1']

            row_ct = sh_demo.max_row
            col_ct = sh_demo.max_column

            max_col_thucdon = sheet.max_column
            #____________________________________________#
            write_it(sheet, sh_demo, row_ct, col_ct)

            def repeat_write_it(sheet_names):
                for name in sheet_names:
                    write_it(workbook[name], wb[name], wb[name].max_row, wb[name].max_column)

            repeat_write_it(sheet_names)

            ##############################################
            workbook.save('ThucDon.xlsx')
            
            pop = self.popup_done()
            pop.mainloop()
        except:
            pop = self.popup_failed()
            pop.mainloop()
            
#############################################################################################

    def button_add_all(self):
        def roundup(n):
            n = math.trunc(n)
            if (n < 5):
                return n
            elif (n % 10 > 5):
                while n % 10 != 0:
                    n+=1
                return n
            elif (n % 10 < 5):
                while n % 10 != 0:
                    n-=1
                return n
            else:
                return n
        try:
            ###############################
            p.save_book_as(file_name=filename_DD01,
                           dest_file_name='demo.xlsx')
            ###############################
            workbook = load_workbook(filename="Thucdon.xlsx")
            switch = 0

            def write_70717(i):
                sheet.cell(row=i,column=4+max_col_thucdon).value = sh_demo.cell(row=i, column=5).value - 1
                sheet.cell(row=i,column=5+max_col_thucdon).value = 0
                sheet.cell(row=i,column=6+max_col_thucdon).value = sh_demo.cell(row=i, column=5).value - 1
                sheet.cell(row=i,column=7+max_col_thucdon).value = 1
                sheet.cell(row=i,column=8+max_col_thucdon).value = sh_demo.cell(row=i, column=5).value - 1
                sheet.cell(row=i-1,column=4+max_col_thucdon).value = 7
                sheet.cell(row=i-1,column=5+max_col_thucdon).value = 0
                sheet.cell(row=i-1,column=6+max_col_thucdon).value = 7
                sheet.cell(row=i-1,column=7+max_col_thucdon).value = 1
                sheet.cell(row=i-1,column=8+max_col_thucdon).value = 7
            
            def write_it(sheet, sh_demo, row_ct, col_ct, count):
                switch = 0
                for i in range(1, row_ct):
                    for j in range(3,6):          
                        sheet.cell(row=i, column=j-1+max_col_thucdon).value = sh_demo.cell(row=i, column=j).value
                        sheet.cell(row=i, column=1+max_col_thucdon).value = sh_demo.cell(row=i, column=1).value
                        if (switch == 1):
                            sheet.cell(row=i, column=j-1+max_col_thucdon).value = ''
                            sheet.cell(row=i, column=j-2+max_col_thucdon).value = ''
                            switch = 0
                        if (type(sheet.cell(row=i, column=j-1+max_col_thucdon).value) is str):
                            if (sheet.cell(row=i, column=j-1+max_col_thucdon).value[:3] == 'SUM'):
                                switch = 1
                                
                switch = 0
                for i in range(1, row_ct):
                    for j in range(1, sh_waste.max_row):
                        if (sh_demo.cell(row=i, column=4).value == sh_waste.cell(row=j, column=2).value):
                            try:
                                if switch == 1:
                                    switch = 0
                                    pass
                                elif sheet.cell(row=i,column=3+max_col_thucdon).value not in special_labels:
                                    sheet.cell(row=i,column=5+max_col_thucdon).value = sh_waste.cell(row=j, column=4).value
                                    sheet.cell(row=i,column=6+max_col_thucdon).value = roundup(((float(sh_waste.cell(row=j, column=4).value) + 100) * sh_demo.cell(row=i, column=5).value)/100)
                                    sheet.cell(row=i,column=7+max_col_thucdon).value = sh_waste.cell(row=j, column=5).value
                                    sheet.cell(row=i,column=8+max_col_thucdon).value = roundup(float(sh_demo.cell(row=i, column=5).value/float(sh_waste.cell(row=j, column=5).value)))
                                elif (sheet.cell(row=i,column=3+max_col_thucdon).value == " Muối (Salt)"):
                                    sheet.cell(row=i,column=5+max_col_thucdon).value = sh_waste.cell(row=j, column=4).value
                                    sheet.cell(row=i,column=6+max_col_thucdon).value = ((float(sh_waste.cell(row=j, column=4).value) + 100) * sh_demo.cell(row=i, column=5).value)/100
                                    sheet.cell(row=i,column=7+max_col_thucdon).value = sh_waste.cell(row=j, column=5).value
                                    sheet.cell(row=i,column=8+max_col_thucdon).value = float(sh_demo.cell(row=i, column=5).value/float(sh_waste.cell(row=j, column=5).value))
                                    if sheet.cell(row=i-1,column=3+max_col_thucdon).value in salty_labels:
                                        if (sheet.cell(row=i-1,column=4+max_col_thucdon).value == 0) and (sheet.cell(row=i,column=3+max_col_thucdon).value > 1):
                                            write_70717(i)
                                    elif sheet.cell(row=i+1,column=3+max_col_thucdon).value in salty_labels:
                                        if (sheet.cell(row=i+1,column=4+max_col_thucdon).value == 0) and (sheet.cell(row=i,column=3+max_col_thucdon).value > 1):
                                            write_70717(i)
                                            switch = 1
                                else:
                                    sheet.cell(row=i,column=5+max_col_thucdon).value = sh_waste.cell(row=j, column=4).value
                                    sheet.cell(row=i,column=6+max_col_thucdon).value = ((float(sh_waste.cell(row=j, column=4).value) + 100) * sh_demo.cell(row=i, column=5).value)/100
                                    sheet.cell(row=i,column=7+max_col_thucdon).value = sh_waste.cell(row=j, column=5).value
                                    sheet.cell(row=i,column=8+max_col_thucdon).value = float(sh_demo.cell(row=i, column=5).value/float(sh_waste.cell(row=j, column=5).value))
                                break
                            except:
                                pass
                            
                    if (sheet.cell(row=i, column=4+max_col_thucdon).value == "[Weight]"):
                        if (count == 0):
                            sheet.cell(row=i, column=4+max_col_thucdon).value = "DD01"
                            sheet.cell(row=i, column=6+max_col_thucdon).value = "DD01"
                            sheet.cell(row=i, column=8+max_col_thucdon).value = "DD01"
                        elif (count == 1):
                            sheet.cell(row=i, column=4+max_col_thucdon).value = "TM01"
                            sheet.cell(row=i, column=6+max_col_thucdon).value = "TM01"
                            sheet.cell(row=i, column=8+max_col_thucdon).value = "TM01"
                        sheet.cell(row=i, column=5+max_col_thucdon).value = "[Waste]"
                        sheet.cell(row=i, column=7+max_col_thucdon).value = "[Tỉ lệ sống/chín]"

                        sheet.cell(row=i+1, column=5+max_col_thucdon).value = "(%)"
                        sheet.cell(row=i+1, column=6+max_col_thucdon).value = "(g)"
                        sheet.cell(row=i+1, column=7+max_col_thucdon).value = "(%)"
                        sheet.cell(row=i+1, column=8+max_col_thucdon).value = "(g)"

                sheet.cell(row=1, column=4+max_col_thucdon).value = "[Weight]"
                sheet.cell(row=1, column=5+max_col_thucdon).value = "[Waste]"
                sheet.cell(row=1, column=6+max_col_thucdon).value = "[Đi chợ]"
                sheet.cell(row=1, column=7+max_col_thucdon).value = "[Tỉ lệ sống/chín]"
                sheet.cell(row=1, column=8+max_col_thucdon).value = "[Phân chia]"

            ################Thứ 2 - Món 1#########################
            sheet = workbook['Thứ 2 - Món 1']
            wb = load_workbook(filename='demo.xlsx')
            sh_demo = wb['Thứ 2 - Món 1']

            row_ct = sh_demo.max_row
            col_ct = sh_demo.max_column

            max_col_thucdon = sheet.max_column
            #____________________________________________#
            write_it(sheet, sh_demo, row_ct, col_ct,0)

            def repeat_write_it(sheet_names):
                for name in sheet_names:
                    write_it(workbook[name], wb[name], wb[name].max_row, wb[name].max_column,0)

            repeat_write_it(sheet_names)
            
            ##############################################
            workbook.save('ThucDon.xlsx')

            ####################################################################################

            p.save_book_as(file_name=filename_TM01,
                           dest_file_name='demo.xlsx')
            ###############################
            workbook = load_workbook(filename="Thucdon.xlsx")
            switch = 0
                            
            ################Thứ 2 - Món 1#########################
            sheet = workbook['Thứ 2 - Món 1']
            wb = load_workbook(filename='demo.xlsx')
            sh_demo = wb['Thứ 2 - Món 1']

            row_ct = sh_demo.max_row
            col_ct = sh_demo.max_column

            max_col_thucdon = sheet.max_column
            #____________________________________________#
            write_it(sheet, sh_demo, row_ct, col_ct,1)

            def repeat_write_it(sheet_names):
                for name in sheet_names:
                    write_it(workbook[name], wb[name], wb[name].max_row, wb[name].max_column,1)

            repeat_write_it(sheet_names)
            
            ##############################################
            workbook.save('ThucDon.xlsx')

            pop = self.popup_done()
            pop.mainloop()
        except:
            pop = self.popup_failed()
            pop.mainloop()

#############################################################################################
    def button_add_DD01(self):
        def roundup(n):
            n = math.trunc(n)
            if (n < 5):
                return n
            elif (n % 10 > 5):
                while n % 10 != 0:
                    n+=1
                return n
            elif (n % 10 < 5):
                while n % 10 != 0:
                    n-=1
                return n
            else:
                return n
        try:
            ###############################
            p.save_book_as(file_name=filename_DD01,
                           dest_file_name='demo.xlsx')
            ###############################
            workbook = load_workbook(filename="Thucdon.xlsx")
            switch = 0

            def write_70717(i):
                sheet.cell(row=i,column=4+max_col_thucdon).value = sh_demo.cell(row=i, column=5).value - 1
                sheet.cell(row=i,column=5+max_col_thucdon).value = 0
                sheet.cell(row=i,column=6+max_col_thucdon).value = sh_demo.cell(row=i, column=5).value - 1
                sheet.cell(row=i,column=7+max_col_thucdon).value = 1
                sheet.cell(row=i,column=8+max_col_thucdon).value = sh_demo.cell(row=i, column=5).value - 1
                sheet.cell(row=i-1,column=4+max_col_thucdon).value = 7
                sheet.cell(row=i-1,column=5+max_col_thucdon).value = 0
                sheet.cell(row=i-1,column=6+max_col_thucdon).value = 7
                sheet.cell(row=i-1,column=7+max_col_thucdon).value = 1
                sheet.cell(row=i-1,column=8+max_col_thucdon).value = 7
            
            def write_it(sheet, sh_demo, row_ct, col_ct, count):
                switch = 0
                for i in range(1, row_ct):
                    for j in range(3,6):          
                        sheet.cell(row=i, column=j-1+max_col_thucdon).value = sh_demo.cell(row=i, column=j).value
                        sheet.cell(row=i, column=1+max_col_thucdon).value = sh_demo.cell(row=i, column=1).value
                        if (switch == 1):
                            sheet.cell(row=i, column=j-1+max_col_thucdon).value = ''
                            sheet.cell(row=i, column=j-2+max_col_thucdon).value = ''
                            switch = 0
                        if (type(sheet.cell(row=i, column=j-1+max_col_thucdon).value) is str):
                            if (sheet.cell(row=i, column=j-1+max_col_thucdon).value[:3] == 'SUM'):
                                switch = 1
                                
                switch = 0
                for i in range(1, row_ct):
                    for j in range(1, sh_waste.max_row):
                        if (sh_demo.cell(row=i, column=4).value == sh_waste.cell(row=j, column=2).value):
                            try:
                                if switch == 1:
                                    switch = 0
                                    pass
                                elif sheet.cell(row=i,column=3+max_col_thucdon).value not in special_labels:
                                    sheet.cell(row=i,column=5+max_col_thucdon).value = sh_waste.cell(row=j, column=4).value
                                    sheet.cell(row=i,column=6+max_col_thucdon).value = roundup(((float(sh_waste.cell(row=j, column=4).value) + 100) * sh_demo.cell(row=i, column=5).value)/100)
                                    sheet.cell(row=i,column=7+max_col_thucdon).value = sh_waste.cell(row=j, column=5).value
                                    sheet.cell(row=i,column=8+max_col_thucdon).value = roundup(float(sh_demo.cell(row=i, column=5).value/float(sh_waste.cell(row=j, column=5).value)))
                                elif (sheet.cell(row=i,column=3+max_col_thucdon).value == " Muối (Salt)"):
                                    sheet.cell(row=i,column=5+max_col_thucdon).value = sh_waste.cell(row=j, column=4).value
                                    sheet.cell(row=i,column=6+max_col_thucdon).value = ((float(sh_waste.cell(row=j, column=4).value) + 100) * sh_demo.cell(row=i, column=5).value)/100
                                    sheet.cell(row=i,column=7+max_col_thucdon).value = sh_waste.cell(row=j, column=5).value
                                    sheet.cell(row=i,column=8+max_col_thucdon).value = float(sh_demo.cell(row=i, column=5).value/float(sh_waste.cell(row=j, column=5).value))
                                    if sheet.cell(row=i-1,column=3+max_col_thucdon).value in salty_labels:
                                        if (sheet.cell(row=i-1,column=4+max_col_thucdon).value == 0) and (sheet.cell(row=i,column=3+max_col_thucdon).value > 1):
                                            write_70717(i)
                                    elif sheet.cell(row=i+1,column=3+max_col_thucdon).value in salty_labels:
                                        if (sheet.cell(row=i+1,column=4+max_col_thucdon).value == 0) and (sheet.cell(row=i,column=3+max_col_thucdon).value > 1):
                                            write_70717(i)
                                            switch = 1
                                else:
                                    sheet.cell(row=i,column=5+max_col_thucdon).value = sh_waste.cell(row=j, column=4).value
                                    sheet.cell(row=i,column=6+max_col_thucdon).value = ((float(sh_waste.cell(row=j, column=4).value) + 100) * sh_demo.cell(row=i, column=5).value)/100
                                    sheet.cell(row=i,column=7+max_col_thucdon).value = sh_waste.cell(row=j, column=5).value
                                    sheet.cell(row=i,column=8+max_col_thucdon).value = float(sh_demo.cell(row=i, column=5).value/float(sh_waste.cell(row=j, column=5).value))
                                break
                            except:
                                pass
                            
                    if (sheet.cell(row=i, column=4+max_col_thucdon).value == "[Weight]"):
                        sheet.cell(row=i, column=4+max_col_thucdon).value = "DD01"
                        sheet.cell(row=i, column=5+max_col_thucdon).value = "[Waste]"
                        sheet.cell(row=i, column=6+max_col_thucdon).value = "DD01"
                        sheet.cell(row=i, column=7+max_col_thucdon).value = "[Tỉ lệ sống/chín]"
                        sheet.cell(row=i, column=8+max_col_thucdon).value = "DD01"

                        sheet.cell(row=i+1, column=5+max_col_thucdon).value = "(%)"
                        sheet.cell(row=i+1, column=6+max_col_thucdon).value = "(g)"
                        sheet.cell(row=i+1, column=7+max_col_thucdon).value = "(%)"
                        sheet.cell(row=i+1, column=8+max_col_thucdon).value = "(g)"

                sheet.cell(row=1, column=4+max_col_thucdon).value = "[Weight]"
                sheet.cell(row=1, column=5+max_col_thucdon).value = "[Waste]"
                sheet.cell(row=1, column=6+max_col_thucdon).value = "[Đi chợ]"
                sheet.cell(row=1, column=7+max_col_thucdon).value = "[Tỉ lệ sống/chín]"
                sheet.cell(row=1, column=8+max_col_thucdon).value = "[Phân chia]"

            ################Thứ 2 - Món 1#########################
            sheet = workbook['Thứ 2 - Món 1']
            wb = load_workbook(filename='demo.xlsx')
            sh_demo = wb['Thứ 2 - Món 1']

            row_ct = sh_demo.max_row
            col_ct = sh_demo.max_column

            max_col_thucdon = sheet.max_column
            #____________________________________________#
            write_it(sheet, sh_demo, row_ct, col_ct,0)

            def repeat_write_it(sheet_names):
                for name in sheet_names:
                    write_it(workbook[name], wb[name], wb[name].max_row, wb[name].max_column,0)

            repeat_write_it(sheet_names)
            
            ##############################################
            workbook.save('ThucDon.xlsx')

            pop = self.popup_done()
            pop.mainloop()
        except:
            pop = self.popup_failed()
            pop.mainloop()

    def button_add_TM01(self):
        def roundup(n):
            n = math.trunc(n)
            if (n < 5):
                return n
            elif (n % 10 > 5):
                while n % 10 != 0:
                    n+=1
                return n
            elif (n % 10 < 5):
                while n % 10 != 0:
                    n-=1
                return n
            else:
                return n
        try:
            ###############################
            p.save_book_as(file_name=filename_DD01,
                           dest_file_name='demo.xlsx')
            ###############################
            workbook = load_workbook(filename="Thucdon.xlsx")
            switch = 0

            def write_70717(i):
                sheet.cell(row=i,column=4+max_col_thucdon).value = sh_demo.cell(row=i, column=5).value - 1
                sheet.cell(row=i,column=5+max_col_thucdon).value = 0
                sheet.cell(row=i,column=6+max_col_thucdon).value = sh_demo.cell(row=i, column=5).value - 1
                sheet.cell(row=i,column=7+max_col_thucdon).value = 1
                sheet.cell(row=i,column=8+max_col_thucdon).value = sh_demo.cell(row=i, column=5).value - 1
                sheet.cell(row=i-1,column=4+max_col_thucdon).value = 7
                sheet.cell(row=i-1,column=5+max_col_thucdon).value = 0
                sheet.cell(row=i-1,column=6+max_col_thucdon).value = 7
                sheet.cell(row=i-1,column=7+max_col_thucdon).value = 1
                sheet.cell(row=i-1,column=8+max_col_thucdon).value = 7
                
            def write_it(sheet, sh_demo, row_ct, col_ct, count):
                switch = 0
                for i in range(1, row_ct):
                    for j in range(3,6):          
                        sheet.cell(row=i, column=j-1+max_col_thucdon).value = sh_demo.cell(row=i, column=j).value
                        sheet.cell(row=i, column=1+max_col_thucdon).value = sh_demo.cell(row=i, column=1).value
                        if (switch == 1):
                            sheet.cell(row=i, column=j-1+max_col_thucdon).value = ''
                            sheet.cell(row=i, column=j-2+max_col_thucdon).value = ''
                            switch = 0
                        if (type(sheet.cell(row=i, column=j-1+max_col_thucdon).value) is str):
                            if (sheet.cell(row=i, column=j-1+max_col_thucdon).value[:3] == 'SUM'):
                                switch = 1
                                
                switch = 0
                for i in range(1, row_ct):
                    for j in range(1, sh_waste.max_row):
                        if (sh_demo.cell(row=i, column=4).value == sh_waste.cell(row=j, column=2).value):
                            try:
                                if switch == 1:
                                    switch = 0
                                    pass
                                elif sheet.cell(row=i,column=3+max_col_thucdon).value not in special_labels:
                                    sheet.cell(row=i,column=5+max_col_thucdon).value = sh_waste.cell(row=j, column=4).value
                                    sheet.cell(row=i,column=6+max_col_thucdon).value = roundup(((float(sh_waste.cell(row=j, column=4).value) + 100) * sh_demo.cell(row=i, column=5).value)/100)
                                    sheet.cell(row=i,column=7+max_col_thucdon).value = sh_waste.cell(row=j, column=5).value
                                    sheet.cell(row=i,column=8+max_col_thucdon).value = roundup(float(sh_demo.cell(row=i, column=5).value/float(sh_waste.cell(row=j, column=5).value)))
                                elif (sheet.cell(row=i,column=3+max_col_thucdon).value == " Muối (Salt)"):
                                    sheet.cell(row=i,column=5+max_col_thucdon).value = sh_waste.cell(row=j, column=4).value
                                    sheet.cell(row=i,column=6+max_col_thucdon).value = ((float(sh_waste.cell(row=j, column=4).value) + 100) * sh_demo.cell(row=i, column=5).value)/100
                                    sheet.cell(row=i,column=7+max_col_thucdon).value = sh_waste.cell(row=j, column=5).value
                                    sheet.cell(row=i,column=8+max_col_thucdon).value = float(sh_demo.cell(row=i, column=5).value/float(sh_waste.cell(row=j, column=5).value))
                                    if sheet.cell(row=i-1,column=3+max_col_thucdon).value in salty_labels:
                                        if (sheet.cell(row=i-1,column=4+max_col_thucdon).value == 0) and (sheet.cell(row=i,column=3+max_col_thucdon).value > 1):
                                            write_70717(i)
                                    elif sheet.cell(row=i+1,column=3+max_col_thucdon).value in salty_labels:
                                        if (sheet.cell(row=i+1,column=4+max_col_thucdon).value == 0) and (sheet.cell(row=i,column=3+max_col_thucdon).value > 1):
                                            write_70717(i)
                                            switch = 1
                                else:
                                    sheet.cell(row=i,column=5+max_col_thucdon).value = sh_waste.cell(row=j, column=4).value
                                    sheet.cell(row=i,column=6+max_col_thucdon).value = ((float(sh_waste.cell(row=j, column=4).value) + 100) * sh_demo.cell(row=i, column=5).value)/100
                                    sheet.cell(row=i,column=7+max_col_thucdon).value = sh_waste.cell(row=j, column=5).value
                                    sheet.cell(row=i,column=8+max_col_thucdon).value = float(sh_demo.cell(row=i, column=5).value/float(sh_waste.cell(row=j, column=5).value))
                                break
                            except:
                                pass
                            
                    if (sheet.cell(row=i, column=4+max_col_thucdon).value == "[Weight]"):
                        sheet.cell(row=i, column=4+max_col_thucdon).value = "TM01"
                        sheet.cell(row=i, column=5+max_col_thucdon).value = "[Waste]"
                        sheet.cell(row=i, column=6+max_col_thucdon).value = "TM01"
                        sheet.cell(row=i, column=7+max_col_thucdon).value = "[Tỉ lệ sống/chín]"
                        sheet.cell(row=i, column=8+max_col_thucdon).value = "DD01"

                        sheet.cell(row=i+1, column=5+max_col_thucdon).value = "(%)"
                        sheet.cell(row=i+1, column=6+max_col_thucdon).value = "(g)"
                        sheet.cell(row=i+1, column=7+max_col_thucdon).value = "(%)"
                        sheet.cell(row=i+1, column=8+max_col_thucdon).value = "(g)"

                sheet.cell(row=1, column=4+max_col_thucdon).value = "[Weight]"
                sheet.cell(row=1, column=5+max_col_thucdon).value = "[Waste]"
                sheet.cell(row=1, column=6+max_col_thucdon).value = "[Đi chợ]"
                sheet.cell(row=1, column=7+max_col_thucdon).value = "[Tỉ lệ sống/chín]"
                sheet.cell(row=1, column=8+max_col_thucdon).value = "[Phân chia]"

            ################Thứ 2 - Món 1#########################
            sheet = workbook['Thứ 2 - Món 1']
            wb = load_workbook(filename='demo.xlsx')
            sh_demo = wb['Thứ 2 - Món 1']

            row_ct = sh_demo.max_row
            col_ct = sh_demo.max_column

            max_col_thucdon = sheet.max_column
            #____________________________________________#
            write_it(sheet, sh_demo, row_ct, col_ct,0)

            def repeat_write_it(sheet_names):
                for name in sheet_names:
                    write_it(workbook[name], wb[name], wb[name].max_row, wb[name].max_column,0)

            repeat_write_it(sheet_names)
            
            ##############################################
            workbook.save('ThucDon.xlsx')

            pop = self.popup_done()
            pop.mainloop()
        except:
            pop = self.popup_failed()
            pop.mainloop()
    
    def on_closing(self, event=0):
        self.destroy()


if __name__ == "__main__":
    app = App()
    app.mainloop()
    
